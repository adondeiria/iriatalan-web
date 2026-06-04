from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

RED = "9E1B1E"; REDHEAD = "B5292C"; GOLD = "C4A06B"; INK = "2B2B2B"; ZEBRA = "F4F4F4"
FONT = "Calibri"

wb = Workbook(); ws = wb.active; ws.title = "Check-up Beneficiarios"
ws.sheet_view.showGridLines = False
for c, w in {"A": 60, "B": 18, "C": 30, "D": 30, "E": 16}.items(): ws.column_dimensions[c].width = w
thin = Side(style="thin", color="D8D8D8"); border = Border(left=thin, right=thin, top=thin, bottom=thin)

def merge(r, text, *, bg=None, fg=INK, sz=11, bold=False, italic=False, h="left", height=None):
    ws.merge_cells(f"A{r}:E{r}"); cell = ws[f"A{r}"]; cell.value = text
    cell.font = Font(name=FONT, size=sz, bold=bold, italic=italic, color=fg)
    cell.alignment = Alignment(horizontal=h, vertical="center", wrap_text=True)
    if bg:
        for col in "ABCDE": ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=bg)
    if height: ws.row_dimensions[r].height = height

def section(r, text): merge(r, text, bg=RED, fg="FFFFFF", sz=13, bold=True, height=26)
def subhead(r, text): merge(r, text, fg=RED, sz=11, bold=True, height=18)

dv_estado = DataValidation(type="list", formula1='"Correcto,Desactualizado,No existe,No aplica,No estoy seguro(a)"', allow_blank=True)
dv_alerta = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
dv_exp = DataValidation(type="list", formula1='"Lo tengo,Me falta,No aplica"', allow_blank=True)
for dv in (dv_estado, dv_alerta, dv_exp): ws.add_data_validation(dv)

r = 1
merge(r, "Check-up Patrimonial de Beneficiarios", bg=RED, fg="FFFFFF", sz=18, bold=True, h="center", height=34); r+=1
merge(r, "Revisa si los beneficiarios en tus seguros, cuentas, retiro, créditos y registros están actualizados y alineados con tu testamento.", bg=GOLD, fg=INK, sz=10, h="center", italic=True, height=30); r+=1
merge(r, "Por Iria Talan · Asesora Patrimonial y de Seguros · Cédula CNSF V388618", fg="7A6A55", sz=9, h="center", height=16); r+=2
subhead(r, "Cómo usar este documento"); r+=1
for t in ["1.  Reúne tus pólizas, contratos, estados de cuenta y documentos de identificación.",
          "2.  En la columna Estado marca cada rubro: Correcto · Desactualizado · No existe · No aplica · No estoy seguro(a).",
          "3.  Si detectas algo, anota el riesgo, la acción a tomar y una fecha compromiso.",
          "4.  Agenda la actualización con tu aseguradora, banco, AFORE o institución correspondiente."]:
    merge(r, t, sz=10, height=15); r+=1
r+=1
section(r, "Datos generales"); r+=1
for label in ["Nombre completo:", "Fecha de revisión:", "Estado civil actual:", "Hijos o dependientes económicos:", "Persona de contacto de confianza:", "Asesor / agente / abogado / notario:"]:
    ws[f"A{r}"].value = label; ws[f"A{r}"].font = Font(name=FONT, size=10, bold=True, color=INK); ws[f"A{r}"].alignment = Alignment(vertical="center")
    ws.merge_cells(f"B{r}:E{r}")
    for col in "BCDE": ws[f"{col}{r}"].fill = PatternFill("solid", fgColor="FFFFFF"); ws[f"{col}{r}"].border = border
    ws.row_dimensions[r].height = 18; r+=1
r+=1

def table_header(r):
    for i, h in enumerate(["Ítem a revisar", "Estado", "Riesgo detectado", "Acción a tomar", "Fecha compromiso"]):
        col = chr(65+i); c = ws[f"{col}{r}"]; c.value = h
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=REDHEAD)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = border
    ws.row_dimensions[r].height = 20; return r+1

def items_table(r, items, dv):
    r = table_header(r)
    for idx, it in enumerate(items):
        ws[f"A{r}"].value = it; ws[f"A{r}"].font = Font(name=FONT, size=10, color=INK)
        fill = PatternFill("solid", fgColor=(ZEBRA if idx%2==0 else "FFFFFF"))
        for col in "ABCDE":
            ws[f"{col}{r}"].border = border; ws[f"{col}{r}"].fill = fill
            ws[f"{col}{r}"].alignment = Alignment(vertical="center", wrap_text=True)
        dv.add(ws[f"B{r}"]); ws.row_dimensions[r].height = 28; r+=1
    return r

def docs(r, items):
    merge(r, "Documentos a tener a la mano:  " + " · ".join(items), fg="7A6A55", sz=9, italic=True); r+=1; return r

# (titulo, items, documentos)
SECS = [
 ("1. Seguros de vida y accidentes", [
   "Tengo localizadas todas mis pólizas vigentes.",
   "Sé quién aparece hoy como beneficiario principal en cada póliza.",
   "Sé quién aparece como beneficiario sustituto o contingente, si aplica.",
   "Los nombres están completos y correctamente escritos.",
   "Los porcentajes de distribución suman 100%.",
   "No dejé expresiones genéricas como “mis hijos” o “mi familia” si la póliza requiere identificación precisa.",
   "Mis beneficiarios siguen siendo las personas que hoy quiero proteger.",
   "No tengo registrada a una expareja o a una persona fallecida.",
   "Si hay menores de edad, ya revisé cómo debe administrarse el recurso y con qué figura legal."],
  ["Carátula de la póliza", "Formato de designación o cambio de beneficiarios", "Identificaciones de los beneficiarios", "CURP y actas de parentesco"]),
 ("2. Planes Personales de Retiro (PPR) y seguros de ahorro", [
   "Sé si tengo un PPR o seguro de ahorro/retiro con una aseguradora (distinto del AFORE).",
   "Sé quién aparece como beneficiario de cada plan.",
   "Los beneficiarios del plan están actualizados.",
   "Conozco las condiciones de disposición y qué pasa con el plan al fallecer."],
  ["Carátula o contrato del PPR / seguro de ahorro", "Designación de beneficiarios"]),
 ("3. Cuentas bancarias, inversiones y plataformas digitales", [
   "Sé en qué bancos tengo cuentas, pagarés o inversiones.",
   "Verifiqué si cada cuenta tiene beneficiarios registrados.",
   "Revisé que los porcentajes estén claros.",
   "Mis beneficiarios conocen la existencia de estas cuentas.",
   "Mis estados de cuenta y contratos están ordenados y localizables.",
   "Ya actualicé beneficiarios después de cambios importantes de vida.",
   "Tengo identificadas mis plataformas digitales / wallets de fondos de inversión (apps, casas de bolsa: p. ej. GBM, Fintual, brokers).",
   "Sé si cada plataforma permite designar beneficiarios y, si lo permite, ya lo hice.",
   "Una persona de confianza sabe que estas inversiones digitales existen y cómo se accede a ellas.",
   "Sé qué pasa con el saldo de cada wallet o plataforma al fallecer (designación, sucesión o el proceso propio de la plataforma)."],
  ["Contratos de cuenta o inversión", "Últimos estados de cuenta", "Accesos y estados de cuenta de plataformas digitales", "Datos de los beneficiarios"]),
 ("4. AFORE, IMSS e ISSSTE", [
   "Sé en qué AFORE estoy registrado(a).",
   "Ya confirmé quién aparece como beneficiario en mi cuenta individual.",
   "Mi pareja, hijos u otros dependientes están correctamente registrados donde corresponde.",
   "Mi estado civil está actualizado en mis expedientes institucionales.",
   "Ya registré a mis beneficiarios en IMSS / ISSSTE.",
   "Mis beneficiarios sabrían a qué institución acudir en caso de fallecimiento.",
   "Tengo localizada mi CURP, NSS y documentos básicos del expediente."],
  ["Estado de cuenta de la AFORE", "CURP", "NSS", "Identificación oficial", "Actas de nacimiento / matrimonio / parentesco"]),
 ("5. Créditos y seguros de saldo deudor", [
   "Sé qué créditos tengo vigentes (hipotecario, automotriz, personal, tarjetas).",
   "Sé si cada crédito incluye un seguro de vida o de saldo deudor que liquida la deuda al fallecer.",
   "Mi familia sabe que esos seguros existen para no pagar la deuda de su bolsillo.",
   "Revisé que el seguro cubra el saldo total y siga vigente.",
   "Tengo localizadas las carátulas de esos seguros y sé a quién avisar."],
  ["Contratos de crédito", "Carátulas del seguro de saldo deudor", "Últimos estados de cuenta"]),
 ("6. Fideicomisos", [
   "Sé si tengo algún fideicomiso constituido (patrimonial, testamentario o de seguro).",
   "Sé quién es el fideicomitente, la fiduciaria y el/los fideicomisario(s) de cada uno.",
   "Los fideicomisarios (beneficiarios) están actualizados.",
   "Las instrucciones del fideicomiso reflejan mi voluntad actual.",
   "El fideicomiso está coordinado con mi testamento y mis pólizas (no se contradicen).",
   "Mi familia sabe que existe y a quién acudir."],
  ["Contrato de fideicomiso", "Datos de la institución fiduciaria", "Identificación de los fideicomisarios"]),
 ("7. Empresas y sucesión societaria", [
   "Sé qué pasa con mi participación (acciones o partes sociales) si fallezco.",
   "Los estatutos o un convenio entre socios prevén la sucesión de mi participación.",
   "Existe un seguro de persona clave o de compra-venta entre socios, si aplica.",
   "Mi familia sabe cómo se valuaría y cobraría mi parte del negocio.",
   "La sucesión de la empresa está coordinada con mi testamento."],
  ["Acta constitutiva y estatutos", "Convenio entre socios", "Pólizas de persona clave / compra-venta"]),
 ("8. Activos y cuentas en el extranjero", [
   "Tengo identificadas mis cuentas, inversiones o propiedades fuera de México.",
   "Sé si esas cuentas tienen beneficiario designado (figuras como “payable/transfer on death”).",
   "Sé si necesito un testamento en el país donde están esos activos.",
   "Mis beneficiarios sabrían cómo y dónde reclamar los activos en el extranjero.",
   "Consideré el efecto fiscal y cambiario de heredar activos en otra moneda o país.",
   "Mis documentos del extranjero están traducidos / apostillados si se requiere."],
  ["Estados de cuenta del extranjero", "Escrituras de propiedades fuera", "Testamento internacional, si aplica"]),
 ("9. Activos digitales y criptomonedas", [
   "Tengo un registro de mis cuentas y activos digitales (correo, redes, criptomonedas, billeteras).",
   "Una persona de confianza sabría cómo acceder a ellos si yo falto (sin exponer mis contraseñas hoy).",
   "Sé si mis criptoactivos tienen forma de heredarse (llaves, custodios o beneficiarios).",
   "Mis suscripciones y pagos automáticos están documentados para cancelarse o transferirse."],
  ["Inventario de cuentas y activos digitales", "Ubicación segura de accesos y llaves"]),
 ("10. Testamento y coordinación patrimonial", [
   "Tengo testamento vigente.",
   "Mi testamento refleja mi situación familiar actual.",
   "Los nombres y relaciones familiares coinciden con los de otros documentos.",
   "Mi testamento está alineado con mis pólizas, cuentas, fideicomisos y empresa.",
   "Mi familia sabe con qué notaría o abogado revisar este tema."],
  ["Testamento más reciente", "Datos de la notaría", "Inventario general de bienes", "Relación de pólizas, cuentas e inversiones"]),
 ("11. Tutela de hijos menores", [
   "Si tengo hijos menores, ya pensé quién sería su tutor o guardián si yo falto.",
   "Esa designación está reflejada por escrito donde corresponde (p. ej. en el testamento).",
   "La persona elegida sabe y aceptó esa responsabilidad.",
   "Separé claramente quién cuidaría al menor de quién administraría su dinero."],
  ["Testamento con designación de tutor", "Acuerdos familiares por escrito"]),
 ("12. Voluntad anticipada y decisiones en caso de incapacidad", [
   "Sé qué es una voluntad anticipada (decisiones médicas si no puedo expresarme).",
   "Tengo, o consideré, un documento de voluntad anticipada conforme a mi entidad.",
   "Designé a alguien que pueda tomar decisiones médicas o financieras si quedo incapacitado.",
   "Mi familia sabe dónde está ese documento y a quién corresponde actuar."],
  ["Documento de voluntad anticipada", "Poder notarial / mandato"]),
]
for title, items, dd in SECS:
    section(r, title); r+=1
    r = items_table(r, items, dv_estado)
    r = docs(r, dd); r+=1

# 13. Señales de alerta (ampliada)
section(r, "13. Señales de alerta — ¿detectas alguno de estos focos rojos?"); r+=1
alertas = [
  "Tengo beneficiarios que ya fallecieron.",
  "Tengo beneficiarios que ya no forman parte de mi círculo cercano.",
  "No sé quién aparece como beneficiario en una o más pólizas.",
  "No sé si mis cuentas bancarias o inversiones tienen beneficiarios.",
  "No sé si mi AFORE está alineada con mis registros institucionales.",
  "No sé si mis créditos tienen un seguro que pague la deuda al fallecer.",
  "Tengo activos en el extranjero sin un plan de sucesión claro.",
  "Tengo menores como beneficiarios directos y no he revisado la administración legal del dinero.",
  "Tengo hijos menores y no he definido quién sería su tutor.",
  "Mi testamento dice una cosa, pero mis contratos o fideicomisos podrían decir otra.",
  "Mi familia no sabría dónde encontrar mis documentos ni a quién llamar.",
]
r = items_table(r, alertas, dv_alerta); r+=1

# 14. Expediente básico (ampliado)
section(r, "14. Expediente básico para la familia"); r+=1
exp = [
  "Lista de seguros vigentes (vida, GMM, saldo deudor).",
  "Lista de bancos, cuentas e inversiones.",
  "Estado de cuenta de AFORE y de PPR.",
  "Relación de créditos vigentes y sus seguros.",
  "Contrato(s) de fideicomiso, si aplica.",
  "Documentos de la empresa y convenio entre socios, si aplica.",
  "Datos de cuentas y propiedades en el extranjero, si aplica.",
  "Inventario de accesos a cuentas y activos digitales.",
  "CURP, NSS, RFC y copia de identificación.",
  "Contacto del asesor de seguros, del banco, del notario o abogado.",
  "Carpeta física o digital con ubicación conocida por una persona de confianza.",
]
r = items_table(r, exp, dv_exp); r+=1

# 15. Frecuencia
section(r, "15. ¿Cada cuándo revisar? Actualiza este check-up cuando ocurra:"); r+=1
for t in ["•  Matrimonio o divorcio", "•  Nacimiento de hijos", "•  Fallecimiento de un beneficiario o heredero",
          "•  Cambio importante de patrimonio", "•  Cambio de empleo o institución de seguridad social",
          "•  Contratación de nuevas pólizas, créditos o apertura de nuevas cuentas",
          "•  Apertura o cierre de cuentas/propiedades en el extranjero"]:
    merge(r, t, sz=10, height=15); r+=1
r+=1

# 16. Cierre
section(r, "Cierre"); r+=1
merge(r, "La protección patrimonial no depende solo de tener productos financieros, sino de tenerlos bien coordinados. Un buen check-up de beneficiarios reduce riesgos, acelera cobros y evita que tu familia enfrente trámites más complejos de lo necesario.", sz=10, italic=True, height=46); r+=1
r+=1
merge(r, "¿Detectaste algo que actualizar? Te ayudo a diseñar y alinear tu estrategia patrimonial.  ·  iriatalan.com.mx", bg=GOLD, fg=INK, sz=10, bold=True, h="center", height=24); r+=1

ws.print_options.horizontalCentered = True
ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
out = r"C:\Users\iriat\Downloads\check-up-patrimonial-beneficiarios.xlsx"
wb.save(out)
print("OK ·", ws.max_row, "filas ·", len(SECS)+4, "secciones · guardado en", out)
